import pandas as pd
import numpy as np
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, flash
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from sklearn.preprocessing import StandardScaler, LabelEncoder
import joblib
import os
from datetime import datetime
import io
from functools import wraps
from flask import send_file
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from datetime import timedelta
import re


# ========== CONFIGURACIÓN DE FLASK ==========
app = Flask(__name__)
#app.secret_key = "clave_super_segura_2025_mejorada"
app.secret_key = os.environ.get('SECRET_KEY', 'clave_super_segura_2025_mejorada')
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

# ========== CONFIGURACIÓN DE POSTGRESQL 1 ==========
#DATABASE_URL = os.environ.get('DATABASE_URL', 
#                               'postgresql://postgres:270225@db:5432/predictor_vuelos')
#app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
#app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
#
#db = SQLAlchemy(app)
 
# ========== CONFIGURACIÓN DE POSTGRESQL ==========
DATABASE_URL = os.environ.get('DATABASE_URL')

# Render usa postgres:// pero SQLAlchemy necesita postgresql://
if DATABASE_URL and DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

# Si no hay DATABASE_URL (desarrollo local), usar la default
if not DATABASE_URL:
    DATABASE_URL = 'postgresql://postgres:270225@localhost:5432/predictor_vuelos'

app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

print("📌 Base de datos configurada:", DATABASE_URL.split('@')[1] if '@' in DATABASE_URL else 'local')

# ========== INICIALIZAR SQLAlchemy AQUÍ (ANTES DE LOS MODELOS) ==========
db = SQLAlchemy(app)
# ========== MODELOS DE BASE DE DATOS ==========
class Usuario(db.Model):
    __tablename__ = 'usuarios'
    
    id = db.Column(db.Integer, primary_key=True)
    usuario = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    contrasena = db.Column(db.String(255), nullable=False)
    fecha_creacion = db.Column(db.DateTime, default=datetime.now)
    activo = db.Column(db.Boolean, default=True)

    predicciones = db.relationship(
        'Prediccion',
        backref='usuario_ref',
        lazy=True,
        cascade='all, delete-orphan'
    )

    def set_password(self, contrasena):
        self.contrasena = generate_password_hash(contrasena)
    
    def check_password(self, contrasena):
        return check_password_hash(self.contrasena, contrasena)


class Prediccion(db.Model):
    __tablename__ = 'predicciones'
    
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=False)
    aerolinea = db.Column(db.String(100), nullable=False)
    origen = db.Column(db.String(10), nullable=False)
    destino = db.Column(db.String(10), nullable=False)
    fecha_viaje = db.Column(db.Date, nullable=False)
    hora_salida = db.Column(db.String(10), nullable=False)
    duracion = db.Column(db.Float, nullable=False)
    escalas = db.Column(db.Integer, nullable=False)
    informacion = db.Column(db.String(100), nullable=False)
    precio_predicho = db.Column(db.Float, nullable=False)
    fecha_prediccion = db.Column(db.DateTime, default=datetime.now)

# ========== VARIABLES GLOBALES ==========
modelo = None
scaler = None
label_encoders = None
features = None
datos_cache = None

# ========== DECORADORES ==========
def login_requerido(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# ========== CARGA DE MODELO 1==========
#def cargar_modelo():
#    """Carga el modelo entrenado"""
#    global modelo, scaler, label_encoders, features
#    
#    if os.path.exists('modelo_vuelos.pkl'):
#        modelo = joblib.load('modelo_vuelos.pkl')
#        scaler = joblib.load('scaler.pkl')
#        label_encoders = joblib.load('label_encoders.pkl')
#        features = joblib.load('features.pkl')
#        return True
#    return False
#def cargar_modelo():
#    """Carga el modelo entrenado"""
#    global modelo, scaler, label_encoders, features
#    
#    archivos_modelo = ['modelo_vuelos.pkl', 'scaler.pkl', 'label_encoders.pkl', 'features.pkl']
#    
#    # Verificar si todos los archivos existen
#    if all(os.path.exists(f) for f in archivos_modelo):
#        try:
#            modelo = joblib.load('modelo_vuelos.pkl')
#            scaler = joblib.load('scaler.pkl')
#            label_encoders = joblib.load('label_encoders.pkl')
#            features = joblib.load('features.pkl')
#            print("✓ Modelo cargado exitosamente")
#            return True
#        except Exception as e:
#            print(f"⚠️ Error cargando modelo: {e}")
#    
#    # Si no existen, entrenar el modelo automáticamente
#    print("⚠️ Modelo no encontrado. Entrenando automáticamente...")
#    try:
#        import training
#        resultado = training.main()
#        if resultado:
#            # Intentar cargar nuevamente
#            modelo = joblib.load('modelo_vuelos.pkl')
#            scaler = joblib.load('scaler.pkl')
#            label_encoders = joblib.load('label_encoders.pkl')
#            features = joblib.load('features.pkl')
#            print("✓ Modelo entrenado y cargado exitosamente")
#            return True
#    except Exception as e:
#        print(f"❌ Error entrenando modelo: {e}")
#    
#    return False

#def cargar_datos_cache():
#    """Carga los datos en caché"""
#    global datos_cache
#    
#    if os.path.exists('datos_vuelos.xlsx'):
#        datos_cache = pd.read_excel('datos_vuelos.xlsx')
#    elif os.path.exists('datos_vuelos_peru.xlsx'):
#        datos_cache = pd.read_excel('datos_vuelos_peru.xlsx')
#    
#    return datos_cache is not None

# ========== CARGA DE MODELO ==========
def cargar_modelo():
    """Carga el modelo entrenado"""
    global modelo, scaler, label_encoders, features
    
    archivos_modelo = ['modelo_vuelos.pkl', 'scaler.pkl', 'label_encoders.pkl', 'features.pkl']
    
    # Verificar si todos los archivos existen
    if all(os.path.exists(f) for f in archivos_modelo):
        try:
            modelo = joblib.load('modelo_vuelos.pkl')
            scaler = joblib.load('scaler.pkl')
            label_encoders = joblib.load('label_encoders.pkl')
            features = joblib.load('features.pkl')
            print("✓ Modelo cargado exitosamente")
            return True
        except Exception as e:
            print(f"⚠️ Error cargando modelo: {e}")
    
    # Si no existen, entrenar automáticamente
    print("⚠️ Modelo no encontrado. Entrenando automáticamente...")
    try:
        # Primero generar datos si no existen
        if not os.path.exists('datos_vuelos.xlsx'):
            print("📊 Generando datos de entrenamiento...")
            import generar_datos
            generar_datos.main()
        
        # Luego entrenar modelo
        print("🤖 Entrenando modelo...")
        import training
        resultado = training.main()
        
        if resultado:
            # Cargar modelo recién entrenado
            modelo = joblib.load('modelo_vuelos.pkl')
            scaler = joblib.load('scaler.pkl')
            label_encoders = joblib.load('label_encoders.pkl')
            features = joblib.load('features.pkl')
            print("✓ Modelo entrenado y cargado exitosamente")
            return True
    except Exception as e:
        print(f"❌ Error entrenando modelo: {e}")
        import traceback
        traceback.print_exc()
    
    return False

def cargar_datos_cache():
    """Carga los datos en caché"""
    global datos_cache
    
    # Intentar cargar datos existentes
    if os.path.exists('datos_vuelos.xlsx'):
        try:
            datos_cache = pd.read_excel('datos_vuelos.xlsx')
            print(f"✓ Datos cargados: {len(datos_cache)} registros")
            return True
        except Exception as e:
            print(f"⚠️ Error cargando datos: {e}")
    
    # Si no existen, generar automáticamente
    print("📊 Generando datos automáticamente...")
    try:
        import generar_datos
        generar_datos.main()
        
        # Intentar cargar nuevamente
        if os.path.exists('datos_vuelos.xlsx'):
            datos_cache = pd.read_excel('datos_vuelos.xlsx')
            print(f"✓ Datos generados y cargados: {len(datos_cache)} registros")
            return True
    except Exception as e:
        print(f"❌ Error generando datos: {e}")
        import traceback
        traceback.print_exc()
    
    return False    

# ========== RUTAS DE AUTENTICACIÓN ==========
@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if request.method == 'POST':
        usuario = request.form.get('usuario')
        email = request.form.get('email')
        contrasena = request.form.get('contrasena')
        confirmar_contrasena = request.form.get('confirmar_contrasena')
        
        # Validaciones
        if not usuario or not email or not contrasena:
            flash('Todos los campos son obligatorios', 'danger')
            return redirect(url_for('registro'))
        
        if len(contrasena) < 6:
            flash('La contraseña debe tener al menos 6 caracteres', 'danger')
            return redirect(url_for('registro'))
        
        if contrasena != confirmar_contrasena:
            flash('Las contraseñas no coinciden', 'danger')
            return redirect(url_for('registro'))
        
        # Verificar si usuario existe
        if Usuario.query.filter_by(usuario=usuario).first():
            flash('El usuario ya existe', 'warning')
            return redirect(url_for('registro'))
        
        if Usuario.query.filter_by(email=email).first():
            flash('El email ya está registrado', 'warning')
            return redirect(url_for('registro'))
        
        # Crear nuevo usuario
        nuevo_usuario = Usuario(usuario=usuario, email=email)
        nuevo_usuario.set_password(contrasena)
        
        db.session.add(nuevo_usuario)
        db.session.commit()
        
        flash('¡Registro exitoso! Ahora puedes iniciar sesión', 'success')
        return redirect(url_for('login'))
    
    return render_template('registro.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        usuario = request.form.get('usuario')
        contrasena = request.form.get('contrasena')
        
        usuario_obj = Usuario.query.filter_by(usuario=usuario).first()
        
        if usuario_obj and usuario_obj.check_password(contrasena) and usuario_obj.activo:
            session['usuario_id'] = usuario_obj.id
            session['usuario'] = usuario_obj.usuario
            session['email'] = usuario_obj.email
            flash(f'¡Bienvenido {usuario}!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Usuario o contraseña incorrectos', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Has cerrado sesión correctamente', 'info')
    return redirect(url_for('login'))

# ========== RUTAS PRINCIPALES ==========
#@app.route('/')
#@login_requerido
#def index():
#    if modelo is None:
#        return render_template('error.html', 
#                             mensaje='Modelo no cargado',
#                             detalle='Por favor ejecuta: python training.py')
#    
#    return render_template('index.html')
@app.route('/')
@login_requerido
def index():
    # Intentar cargar el modelo si no está cargado
    if modelo is None:
        print("⚠️ Modelo no cargado, intentando cargar...")
        if not cargar_modelo():
            print("⚠️ No se pudo cargar el modelo")
            # No mostrar error, la app puede seguir funcionando
    
    return render_template('index.html')

@app.route('/dashboard')
@login_requerido
def dashboard():
    usuario_id = session.get('usuario_id')
    predicciones = Prediccion.query.filter_by(usuario_id=usuario_id).all()
    
    # Estadísticas
    total_predicciones = len(predicciones)
    if predicciones:
        precios = [p.precio_predicho for p in predicciones]
        precio_promedio = np.mean(precios)
        precio_min = min(precios)
        precio_max = max(precios)
        
        # Rutas más consultadas
        rutas = {}
        for p in predicciones:
            ruta = f"{p.origen}-{p.destino}"
            rutas[ruta] = rutas.get(ruta, 0) + 1
    else:
        precio_promedio = precio_min = precio_max = 0
        rutas = {}
    
    return render_template('dashboard.html',
                         total_predicciones=total_predicciones,
                         precio_promedio=precio_promedio,
                         precio_min=precio_min,
                         precio_max=precio_max,
                         rutas_top=dict(sorted(rutas.items(), 
                                             key=lambda x: x[1], 
                                             reverse=True)[:5]))

@app.route('/historial')
@login_requerido
def historial():
    usuario_id = session.get('usuario_id')
    predicciones = Prediccion.query.filter_by(usuario_id=usuario_id)\
                                   .order_by(Prediccion.fecha_prediccion.desc())\
                                   .all()
    return render_template('historial.html', predicciones=predicciones)

# ========== RUTAS DE API ==========
@app.route('/api/datos', methods=['GET'])
@login_requerido
def obtener_datos():
    if datos_cache is None:
        return jsonify({'error': 'Datos no disponibles'}), 500
    
    try:
        return jsonify({
            'aerolineas': sorted(datos_cache['Aerolínea'].unique().tolist()),
            'origenes': sorted(datos_cache['Origen'].unique().tolist()),
            'destinos': sorted(datos_cache['Destino'].unique().tolist()),
            'duraciones': sorted(datos_cache['Duración'].unique().tolist()),
            'escalas': sorted(datos_cache['Total_de_escalas'].unique().tolist()),
            'informaciones': sorted(datos_cache['Información_adicional'].unique().tolist())
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/predecir', methods=['POST'])
@login_requerido
def predecir():
    if modelo is None:
        return jsonify({'error': 'Modelo no cargado'}), 500
    
    try:
        datos = request.json
        usuario_id = session.get('usuario_id')
        
        # Validación
        if datos['origen'] == datos['destino']:
            return jsonify({
                'exito': False,
                'error': 'El origen y destino no pueden ser iguales'
            }), 400
        
        # Procesar fecha
        fecha = pd.to_datetime(datos['fecha'])
        fecha_min = pd.to_datetime(datos_cache['Fecha_del_viaje'].min())
        dias_desde_inicio = (fecha - fecha_min).days
        
        # Crear entrada
        entrada = pd.DataFrame({
            'Aerolínea': [label_encoders['Aerolínea'].transform([datos['aerolinea']])[0]],
            'Día_semana': [fecha.dayofweek],
            'Mes': [fecha.month],
            'Trimestre': [fecha.quarter],
            'Es_fin_de_semana': [1 if fecha.weekday() >= 5 else 0],
            'Origen': [label_encoders['Origen'].transform([datos['origen']])[0]],
            'Destino': [label_encoders['Destino'].transform([datos['destino']])[0]],
            'Duración': [float(datos['duracion'])],
            'Total_de_escalas': [int(datos['escalas'])],
            'Información_adicional': [label_encoders['Información_adicional'].transform([datos['informacion']])[0]],
            'Hora_salida_num': [int(datos['hora_salida'].split(':')[0])],
            'Minuto_salida': [int(datos['hora_salida'].split(':')[1])],
            'Días_desde_inicio': [dias_desde_inicio],
            'Longitud_ruta': [len(f"{datos['origen']}-{datos['destino']}")]
        })
        
        # Predicción
        entrada_scaled = scaler.transform(entrada[features])
        precio_predicho = float(modelo.predict(entrada_scaled)[0])
        precio_predicho = max(150, round(precio_predicho, 2))
        
        # Guardar en base de datos
        prediccion = Prediccion(
            usuario_id=usuario_id,
            aerolinea=datos['aerolinea'],
            origen=datos['origen'],
            destino=datos['destino'],
            fecha_viaje=fecha.date(),
            hora_salida=datos['hora_salida'],
            duracion=float(datos['duracion']),
            escalas=int(datos['escalas']),
            informacion=datos['informacion'],
            precio_predicho=precio_predicho
        )
        
        db.session.add(prediccion)
        db.session.commit()
        
        return jsonify({
            'exito': True,
            'precio': precio_predicho,
            'fecha': datos['fecha'],
            'aerolinea': datos['aerolinea'],
            'ruta': f"{datos['origen']} → {datos['destino']}"
        })
    
    except Exception as e:
        return jsonify({'exito': False, 'error': str(e)}), 400

@app.route('/api/historial-json', methods=['GET'])
@login_requerido
def historial_json():
    usuario_id = session.get('usuario_id')
    predicciones = Prediccion.query.filter_by(usuario_id=usuario_id)\
                                   .order_by(Prediccion.fecha_prediccion.desc())\
                                   .limit(10).all()
    
    return jsonify([{
        'id': p.id,
        'aerolinea': p.aerolinea,
        'ruta': f"{p.origen}-{p.destino}",
        'fecha': p.fecha_viaje.strftime('%Y-%m-%d'),
        'precio': p.precio_predicho,
        'hora': p.fecha_prediccion.strftime('%H:%M:%S')
    } for p in predicciones])

@app.route('/api/estadisticas', methods=['GET'])
@login_requerido
def estadisticas():
    if datos_cache is None:
        return jsonify({'error': 'Datos no disponibles'}), 500
    
    try:
        return jsonify({
            'total_registros': len(datos_cache),
            'precio_min': float(datos_cache['Precio (S/)'].min()),
            'precio_max': float(datos_cache['Precio (S/)'].max()),
            'precio_promedio': float(datos_cache['Precio (S/)'].mean()),
            'precio_mediana': float(datos_cache['Precio (S/)'].median()),
            'desviacion_estandar': float(datos_cache['Precio (S/)'].std()),
            'duracion_promedio': float(datos_cache['Duración'].mean()),
            'escalas_promedio': float(datos_cache['Total_de_escalas'].mean())
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/perfil', methods=['GET'])
@login_requerido
def perfil():
    usuario_id = session.get('usuario_id')
    usuario = Usuario.query.get(usuario_id)
    predicciones = Prediccion.query.filter_by(usuario_id=usuario_id).count()
    
    return jsonify({
        'usuario': usuario.usuario,
        'email': usuario.email,
        'fecha_creacion': usuario.fecha_creacion.strftime('%Y-%m-%d'),
        'total_predicciones': predicciones
    })


@app.route('/api/perfil/actualizar', methods=['PUT'])
@login_requerido
def actualizar_perfil():
    """Actualiza los datos del perfil del usuario"""
    usuario_id = session.get('usuario_id')
    usuario = Usuario.query.get(usuario_id)
    
    try:
        datos = request.json
        
        # Validar datos
        nuevo_usuario = datos.get('usuario', '').strip()
        nuevo_email = datos.get('email', '').strip()
        
        if not nuevo_usuario or not nuevo_email:
            return jsonify({'exito': False, 'error': 'Todos los campos son obligatorios'}), 400
        
        # Verificar si el nuevo usuario ya existe (excepto el actual)
        if nuevo_usuario != usuario.usuario:
            existe_usuario = Usuario.query.filter_by(usuario=nuevo_usuario).first()
            if existe_usuario:
                return jsonify({'exito': False, 'error': 'El nombre de usuario ya está en uso'}), 400
        
        # Verificar si el nuevo email ya existe (excepto el actual)
        if nuevo_email != usuario.email:
            existe_email = Usuario.query.filter_by(email=nuevo_email).first()
            if existe_email:
                return jsonify({'exito': False, 'error': 'El email ya está registrado'}), 400
        
        # Actualizar datos
        usuario.usuario = nuevo_usuario
        usuario.email = nuevo_email
        
        db.session.commit()
        
        # Actualizar sesión
        session['usuario'] = nuevo_usuario
        session['email'] = nuevo_email
        
        return jsonify({
            'exito': True,
            'mensaje': 'Perfil actualizado exitosamente',
            'usuario': nuevo_usuario,
            'email': nuevo_email
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'exito': False, 'error': str(e)}), 500

@app.route('/api/perfil/cambiar-contrasena', methods=['PUT'])
@login_requerido
def cambiar_contrasena():
    """Cambia la contraseña del usuario"""
    usuario_id = session.get('usuario_id')
    usuario = Usuario.query.get(usuario_id)
    
    try:
        datos = request.json
        
        contrasena_actual = datos.get('contrasena_actual', '')
        nueva_contrasena = datos.get('nueva_contrasena', '')
        confirmar_contrasena = datos.get('confirmar_contrasena', '')
        
        # Validaciones
        if not contrasena_actual or not nueva_contrasena or not confirmar_contrasena:
            return jsonify({'exito': False, 'error': 'Todos los campos son obligatorios'}), 400
        
        # Verificar contraseña actual
        if not usuario.check_password(contrasena_actual):
            return jsonify({'exito': False, 'error': 'La contraseña actual es incorrecta'}), 400
        
        # Verificar longitud de nueva contraseña
        if len(nueva_contrasena) < 6:
            return jsonify({'exito': False, 'error': 'La nueva contraseña debe tener al menos 6 caracteres'}), 400
        
        # Verificar que las contraseñas coincidan
        if nueva_contrasena != confirmar_contrasena:
            return jsonify({'exito': False, 'error': 'Las contraseñas no coinciden'}), 400
        
        # Actualizar contraseña
        usuario.set_password(nueva_contrasena)
        db.session.commit()
        
        return jsonify({
            'exito': True,
            'mensaje': 'Contraseña cambiada exitosamente'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'exito': False, 'error': str(e)}), 500
        # ========== RUTAS DE EXPORTACIÓN ==========

@app.route('/api/historial/exportar-excel', methods=['GET'])
@login_requerido
def exportar_excel():
    """Exporta el historial de predicciones a Excel"""
    usuario_id = session.get('usuario_id')
    usuario = Usuario.query.get(usuario_id)
    predicciones = Prediccion.query.filter_by(usuario_id=usuario_id)\
                                   .order_by(Prediccion.fecha_prediccion.desc())\
                                   .all()
    
    if not predicciones:
        return jsonify({'error': 'No hay predicciones para exportar'}), 404
    
    try:
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial de Predicciones"
        
        # Estilos
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:H1')
        titulo = ws['A1']
        titulo.value = f"Historial de Predicciones - {usuario.usuario}"
        titulo.font = Font(bold=True, size=16, color="667EEA")
        titulo.alignment = Alignment(horizontal='center', vertical='center')
        
        # Información del usuario
        ws.merge_cells('A2:H2')
        info = ws['A2']
        info.value = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total: {len(predicciones)} predicciones"
        info.font = Font(size=10, italic=True)
        info.alignment = Alignment(horizontal='center')
        
        # Espacio
        ws.append([])
        
        # Encabezados
        headers = ['#', 'Fecha Viaje', 'Aerolínea', 'Origen', 'Destino', 'Duración (h)', 'Escalas', 'Precio (S/)']
        ws.append(headers)
        
        # Aplicar estilo a encabezados
        for cell in ws[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Datos
        for idx, pred in enumerate(predicciones, 1):
            ws.append([
                idx,
                pred.fecha_viaje.strftime('%Y-%m-%d'),
                pred.aerolinea,
                pred.origen,
                pred.destino,
                pred.duracion,
                pred.escalas,
                pred.precio_predicho
            ])
            
            # Aplicar bordes a todas las celdas
            for cell in ws[ws.max_row]:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar anchos de columna
        column_widths = [5, 15, 20, 10, 10, 12, 10, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Estadísticas al final
        ws.append([])
        precios = [p.precio_predicho for p in predicciones]
        ws.append(['ESTADÍSTICAS', '', '', '', '', '', '', ''])
        ws.append(['Precio Promedio:', '', '', '', '', '', '', f"S/ {sum(precios)/len(precios):.2f}"])
        ws.append(['Precio Mínimo:', '', '', '', '', '', '', f"S/ {min(precios):.2f}"])
        ws.append(['Precio Máximo:', '', '', '', '', '', '', f"S/ {max(precios):.2f}"])
        
        # Estilo para estadísticas
        for row in range(ws.max_row - 3, ws.max_row + 1):
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'H{row}'].font = Font(bold=True, color="667EEA")
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nombre del archivo
        filename = f"historial_{usuario.usuario}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/historial/exportar-pdf', methods=['GET'])
@login_requerido
def exportar_pdf():
    """Exporta el historial de predicciones a PDF"""
    usuario_id = session.get('usuario_id')
    usuario = Usuario.query.get(usuario_id)
    predicciones = Prediccion.query.filter_by(usuario_id=usuario_id)\
                                   .order_by(Prediccion.fecha_prediccion.desc())\
                                   .all()
    
    if not predicciones:
        return jsonify({'error': 'No hay predicciones para exportar'}), 404
    
    try:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#667EEA'),
            spaceAfter=12,
            alignment=1  # Centrado
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            spaceAfter=20,
            alignment=1
        )
        
        # Título
        titulo = Paragraph(f"<b>Historial de Predicciones</b><br/>{usuario.usuario}", title_style)
        elements.append(titulo)
        
        # Información
        info_text = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total: {len(predicciones)} predicciones"
        info = Paragraph(info_text, subtitle_style)
        elements.append(info)
        elements.append(Spacer(1, 0.2*inch))
        
        # Tabla de datos
        data = [['#', 'Fecha', 'Aerolínea', 'Ruta', 'Duración', 'Escalas', 'Precio']]
        
        for idx, pred in enumerate(predicciones, 1):
            data.append([
                str(idx),
                pred.fecha_viaje.strftime('%Y-%m-%d'),
                pred.aerolinea[:15],  # Truncar si es muy largo
                f"{pred.origen}-{pred.destino}",
                f"{pred.duracion}h",
                str(pred.escalas),
                f"S/ {pred.precio_predicho:.2f}"
            ])
        
        # Crear tabla
        table = Table(data, colWidths=[0.5*inch, 1*inch, 1.5*inch, 1*inch, 0.8*inch, 0.7*inch, 1*inch])
        
        # Estilo de tabla
        table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667EEA')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            
            # Contenido
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F7F7')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Estadísticas
        precios = [p.precio_predicho for p in predicciones]
        stats_data = [
            ['ESTADÍSTICAS', ''],
            ['Precio Promedio:', f"S/ {sum(precios)/len(precios):.2f}"],
            ['Precio Mínimo:', f"S/ {min(precios):.2f}"],
            ['Precio Máximo:', f"S/ {max(precios):.2f}"],
            ['Duración Promedio:', f"{sum(p.duracion for p in predicciones)/len(predicciones):.1f}h"],
            ['Total de Escalas:', f"{sum(p.escalas for p in predicciones)}"]
        ]
        
        stats_table = Table(stats_data, colWidths=[2.5*inch, 1.5*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667EEA')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 1), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        elements.append(stats_table)
        
        # Pie de página
        elements.append(Spacer(1, 0.3*inch))
        footer = Paragraph(
            f"<i>Documento generado por AeroPredict © {datetime.now().year}</i>",
            ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=1)
        )
        elements.append(footer)
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Nombre del archivo
        filename = f"historial_{usuario.usuario}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ========== RUTA DEL CHAT BOT  ==========
@app.route('/api/chat-bot', methods=['POST'])
@login_requerido
def chat_bot():
    """Procesa mensajes del chat bot con IA conversacional mejorada"""
    try:
        data = request.json
        mensaje_usuario = data.get('mensaje', '')
        contexto = data.get('contexto', {})
        conversacion = data.get('conversacion', [])
        
        if not mensaje_usuario:
            return jsonify({'exito': False, 'error': 'Mensaje vacío'}), 400
        
        # Generar respuesta inteligente con contexto conversacional
        respuesta, cerrar_chat = generar_respuesta_bot_mejorada(mensaje_usuario, contexto, conversacion)
        
        return jsonify({
            'exito': True,
            'respuesta': respuesta,
            'cerrar_chat': cerrar_chat  # Nuevo: indica si debe cerrarse automáticamente
        })
    
    except Exception as e:
        print(f"Error en chat_bot: {str(e)}")
        return jsonify({
            'exito': False,
            'error': 'Error procesando mensaje'
        }), 500


def detectar_intencion(mensaje):
    """Detecta la intención del usuario usando IA"""
    mensaje_lower = mensaje.lower()
    
    # Saludos
    saludos = ['hola', 'hey', 'buenos dias', 'buenas tardes', 'buenas noches', 'que tal', 'como estas', 'como estás', 'saludos']
    if any(saludo in mensaje_lower for saludo in saludos):
        return 'saludo'
    
    # Despedidas y agradecimientos
    despedidas = ['gracias', 'muchas gracias', 'perfecto', 'excelente', 'ok', 'vale', 'adios', 'adiós', 'chao', 'bye', 'hasta luego', 'nos vemos', 'listo', 'entendido', 'ya esta', 'ya está']
    if any(despedida in mensaje_lower for despedida in despedidas):
        return 'despedida'
    
    # Análisis de predicción
    if any(palabra in mensaje_lower for palabra in ['analiza', 'analizar', 'última', 'ultima', 'predicción', 'prediccion', 'resultado', 'mi vuelo', 'mi precio']):
        return 'analizar_prediccion'
    
    # Ayuda con predicción
    if any(palabra in mensaje_lower for palabra in ['ayuda', 'ayudar', 'ayudame', 'ayúdame', 'necesito', 'quiero', 'puedes']):
        if any(palabra in mensaje_lower for palabra in ['predicción', 'prediccion', 'precio', 'vuelo', 'viajar', 'comprar']):
            return 'ayuda_prediccion'
    
    # Temporada
    if 'temporada' in mensaje_lower:
        return 'temporada'
    
    # Comparar aerolíneas
    if 'compar' in mensaje_lower and ('aerolínea' in mensaje_lower or 'aerolinea' in mensaje_lower):
        return 'comparar_aerolineas'
    
    # Cuándo comprar
    if any(palabra in mensaje_lower for palabra in ['cuándo', 'cuando', 'mejor momento', 'cuando comprar', 'cuándo comprar']):
        return 'cuando_comprar'
    
    # Escalas
    if 'escala' in mensaje_lower:
        return 'escalas'
    
    # Días
    if any(palabra in mensaje_lower for palabra in ['día', 'dia', 'mejor dia', 'mejor día']):
        return 'dias_semana'
    
    # Consejos
    if any(palabra in mensaje_lower for palabra in ['consejo', 'tip', 'recomendación', 'recomendacion', 'sugerencia']):
        return 'consejos'
    
    # Pregunta genérica
    return 'generico'


def generar_respuesta_bot_mejorada(mensaje, contexto, conversacion):
    """Genera respuestas inteligentes con IA conversacional mejorada"""
    
    intencion = detectar_intencion(mensaje)
    cerrar_chat = False  # Por defecto no se cierra
    
    ultima_pred = contexto.get('ultimaPrediccion') if contexto else None
    stats = contexto.get('estadisticas') if contexto else None
    
    # ==================== SALUDOS ====================
    if intencion == 'saludo':
        if ultima_pred:
            return f"""👋 <strong>¡Hola! Encantado de ayudarte</strong><br><br>
Veo que acabas de hacer una predicción para <strong>{ultima_pred.get('ruta', '')}</strong> con un precio de <strong>S/ {ultima_pred.get('precio', 0):.2f}</strong>.<br><br>

¿Quieres que te ayude con alguna de estas cosas?<br><br>

<div style="background: #f0f7ff; padding: 12px; border-radius: 8px; margin: 10px 0;">
📊 <strong>"Analiza mi predicción"</strong> - Te doy un análisis completo<br><br>
🏖️ <strong>"¿Es temporada alta?"</strong> - Te digo si es buen momento<br><br>
💡 <strong>"Dame consejos"</strong> - Tips para ahorrar más<br><br>
✈️ <strong>"Compara aerolíneas"</strong> - Veo si hay mejores opciones
</div>

¿En qué te ayudo? 😊""", False
        else:
            return """👋 <strong>¡Hola! ¿Cómo estás?</strong><br><br>
Soy tu asistente inteligente de vuelos. Puedo ayudarte a:<br><br>

✅ Analizar precios de vuelos<br>
✅ Identificar la mejor temporada para viajar<br>
✅ Comparar aerolíneas<br>
✅ Darte consejos para ahorrar<br><br>

<div style="background: #fff3cd; padding: 12px; border-radius: 8px; border-left: 4px solid #ffc107;">
💡 <strong>Tip:</strong> Primero haz una predicción arriba, y luego puedo darte un análisis completo personalizado.
</div>

¿En qué te puedo ayudar? 😊""", False
    
    # ==================== DESPEDIDAS Y AGRADECIMIENTOS ====================
    elif intencion == 'despedida':
        mensaje_lower = mensaje.lower()
        
        # Detectar si es agradecimiento
        if any(palabra in mensaje_lower for palabra in ['gracias', 'thank', 'excelente', 'perfecto', 'genial']):
            cerrar_chat = True  # Se cerrará automáticamente
            return """😊 <strong>¡De nada! Ha sido un placer ayudarte</strong><br><br>

Recuerda:<br>
✅ Compra con 30-45 días de anticipación<br>
✅ Los martes son los mejores días<br>
✅ Usa modo incógnito siempre<br><br>

<div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 12px; border-radius: 8px;">
💡 Vuelve cuando quieras, ¡estoy aquí para ayudarte! ✈️
</div>

<em style="font-size: 11px; color: #999;">Este chat se cerrará en 2 segundos...</em>""", True
        
        # Despedida normal
        else:
            cerrar_chat = True
            return """👋 <strong>¡Hasta luego!</strong><br><br>

Fue un gusto ayudarte. Recuerda que puedes volver cuando quieras.<br><br>

<strong>¡Buen viaje! ✈️</strong><br><br>

<em style="font-size: 11px; color: #999;">Este chat se cerrará en 2 segundos...</em>""", True
    
    # ==================== AYUDA CON PREDICCIÓN ====================
    elif intencion == 'ayuda_prediccion':
        if ultima_pred:
            # Ya tiene una predicción, ofrecer análisis
            return f"""🤝 <strong>¡Claro que sí! Te ayudo con tu predicción</strong><br><br>

Veo que buscas viajar de <strong>{ultima_pred.get('ruta', '')}</strong> con <strong>{ultima_pred.get('aerolinea', '')}</strong>.<br><br>

Te voy a dar un análisis completo ahora mismo...<br><br>

<em style="font-size: 11px; color: #999;">Analizando datos...</em>""", False
            # Luego automáticamente seguirá con el análisis
        else:
            return """🤝 <strong>¡Por supuesto! Te ayudo con tu predicción</strong><br><br>

Para poder ayudarte mejor, primero necesito que hagas una predicción:<br><br>

<div style="background: #f0f7ff; padding: 12px; border-radius: 8px; border-left: 4px solid #667eea;">
<strong>Pasos:</strong><br>
1️⃣ Usa el formulario de arriba<br>
2️⃣ Selecciona tu ruta y fecha<br>
3️⃣ Haz clic en "Calcular Precio"<br>
4️⃣ Vuelve aquí y te doy un análisis completo 📊
</div>

Una vez tengas tu predicción, puedo decirte:<br>
• Si el precio está alto o bajo<br>
• Si es buen momento para comprar<br>
• Qué aerolínea conviene más<br>
• Tips para ahorrar más dinero<br><br>

¿Quieres que te explique algo más mientras tanto? 😊""", False
    
    # ==================== ANÁLISIS DE PREDICCIÓN ====================
    elif intencion == 'analizar_prediccion':
        if not ultima_pred:
            return """❌ <strong>Aún no tienes predicciones</strong><br><br>
            
Para que pueda analizar tu vuelo, primero necesitas hacer una predicción usando el formulario de arriba.<br><br>

<div style="background: #fff3cd; padding: 12px; border-radius: 8px; border-left: 4px solid #ffc107;">
<strong>💡 Cómo hacerlo:</strong><br>
1. Completa el formulario de "Predicción de Precios"<br>
2. Haz clic en "Calcular Precio"<br>
3. Vuelve aquí y pídeme el análisis
</div>

¿Necesitas ayuda con algo más? 😊""", False
        
        precio = ultima_pred.get('precio', 0)
        ruta = ultima_pred.get('ruta', '')
        aerolinea = ultima_pred.get('aerolinea', '')
        fecha = ultima_pred.get('fecha', '')
        
        # Análisis de precio
        nivel_precio = "NORMAL"
        emoji_precio = "💰"
        comparacion = ""
        recomendacion_precio = ""
        
        if stats:
            promedio = stats.get('precio_promedio', 0)
            if promedio > 0:
                diferencia_pct = ((precio / promedio) - 1) * 100
                
                if diferencia_pct < -15:
                    nivel_precio = "EXCELENTE"
                    emoji_precio = "✅"
                    comparacion = f"{abs(diferencia_pct):.1f}% más barato que el promedio"
                    recomendacion_precio = "¡Este es un precio excelente! Te recomiendo comprar pronto antes de que suba."
                elif diferencia_pct > 15:
                    nivel_precio = "ELEVADO"
                    emoji_precio = "⚠️"
                    comparacion = f"{diferencia_pct:.1f}% más caro que el promedio"
                    recomendacion_precio = "El precio está alto. Considera buscar otras fechas u aerolíneas."
                else:
                    comparacion = f"Diferencia: {diferencia_pct:+.1f}% vs promedio"
                    recomendacion_precio = "El precio está en el rango esperado. Es una opción razonable."
        
        # Análisis de temporada y anticipación
        try:
            fecha_obj = datetime.strptime(fecha, '%Y-%m-%d')
            mes_viaje = fecha_obj.month
            dia_semana = fecha_obj.strftime('%A')
            
            # Traducir día
            dias_es = {'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles', 
                      'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sábado', 'Sunday': 'Domingo'}
            dia_viaje = dias_es.get(dia_semana, dia_semana)
            
            # Temporada
            if mes_viaje in [12, 1, 2, 7, 8]:
                temporada = "ALTA"
                emoji_temp = "🏖️"
                msg_temporada = "Estás viajando en temporada alta. Los precios suelen ser 30-50% más altos."
                color_temp = "#ff4757"
            elif mes_viaje in [6, 9, 10, 11]:
                temporada = "MEDIA"
                emoji_temp = "📊"
                msg_temporada = "Temporada media. Los precios son moderados."
                color_temp = "#ffa502"
            else:
                temporada = "BAJA"
                emoji_temp = "✨"
                msg_temporada = "¡Excelente elección! Temporada baja significa mejores precios."
                color_temp = "#26de81"
            
            # Anticipación
            dias_anticipacion = (fecha_obj - datetime.now()).days
            
            if dias_anticipacion < 0:
                anticipacion_msg = "La fecha ya pasó"
                anticipacion_color = "#999"
                anticipacion_consejo = ""
            elif dias_anticipacion < 15:
                anticipacion_msg = f"Faltan {dias_anticipacion} días - ¡URGENTE!"
                anticipacion_color = "#ff4757"
                anticipacion_consejo = "🚨 ¡Compra HOY! Los precios suben mucho cerca de la fecha de viaje."
            elif dias_anticipacion <= 45:
                anticipacion_msg = f"Faltan {dias_anticipacion} días - MOMENTO ÓPTIMO"
                anticipacion_color = "#26de81"
                anticipacion_consejo = "✅ ¡Perfecto! Estás en la ventana ideal de compra (15-45 días antes)."
            elif dias_anticipacion <= 60:
                anticipacion_msg = f"Faltan {dias_anticipacion} días"
                anticipacion_color = "#667eea"
                anticipacion_consejo = "📅 Buen momento para empezar a monitorear ofertas."
            else:
                anticipacion_msg = f"Faltan {dias_anticipacion} días"
                anticipacion_color = "#ffa502"
                anticipacion_consejo = "⏰ Es muy pronto. Espera 2-3 semanas más para precios más estables."
            
            # Análisis del día de viaje
            dias_baratos = ['Martes', 'Miércoles']
            dias_caros = ['Viernes', 'Domingo']
            
            if dia_viaje in dias_baratos:
                dia_msg = f"✅ ¡Excelente! {dia_viaje} es uno de los días más baratos para viajar."
            elif dia_viaje in dias_caros:
                dia_msg = f"⚠️ {dia_viaje} suele ser más caro. Podrías ahorrar 20-30% viajando martes o miércoles."
            else:
                dia_msg = f"💰 {dia_viaje} tiene precios moderados."
                
        except:
            temporada = "N/A"
            emoji_temp = "❓"
            msg_temporada = ""
            color_temp = "#999"
            anticipacion_msg = "No disponible"
            anticipacion_color = "#999"
            anticipacion_consejo = ""
            dia_msg = ""
        
        # Respuesta completa con análisis detallado
        return f"""📊 <strong>Análisis Completo de tu Vuelo</strong><br><br>

<div style="background: #f8f9fa; padding: 15px; border-radius: 10px; border-left: 4px solid #667eea; margin: 10px 0;">
<strong>📍 Ruta:</strong> {ruta}<br>
<strong>✈️ Aerolínea:</strong> {aerolinea}<br>
<strong>📅 Fecha:</strong> {fecha} ({dia_viaje})<br>
<strong>⏰ Anticipación:</strong> {anticipacion_msg}
</div>

<div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; border-radius: 12px; margin: 15px 0;">
<div style="font-size: 28px; font-weight: bold; margin-bottom: 10px;">
{emoji_precio} S/ {precio:.2f}
</div>
<div style="font-size: 16px; margin-bottom: 5px;">
Estado: <strong>{nivel_precio}</strong>
</div>
<div style="font-size: 13px; opacity: 0.9;">
{comparacion}
</div>
</div>

<div style="background: {color_temp}; color: white; padding: 15px; border-radius: 10px; margin: 10px 0;">
<strong>{emoji_temp} TEMPORADA {temporada}</strong><br>
{msg_temporada}
</div>

<div style="background: {anticipacion_color}; color: white; padding: 15px; border-radius: 10px; margin: 10px 0;">
<strong>⏰ ANTICIPACIÓN</strong><br>
{anticipacion_consejo}
</div>

<div style="background: #f0f7ff; padding: 15px; border-radius: 10px; border-left: 4px solid #667eea; margin: 15px 0;">
<strong>📅 DÍA DE VIAJE</strong><br>
{dia_msg}
</div>

<div style="background: #fff3cd; padding: 15px; border-radius: 10px; border-left: 4px solid #ffc107; margin: 15px 0;">
<strong>💡 MI RECOMENDACIÓN</strong><br>
{recomendacion_precio}<br><br>

<strong>Consejos adicionales:</strong><br>
• Compara con otras aerolíneas antes de decidir<br>
• Usa modo incógnito para buscar<br>
• Configura alertas de precio<br>
• Considera volar martes o miércoles si puedes
</div>

¿Necesitas más información o tienes alguna pregunta? 😊""", False
    
    # ==================== OTRAS INTENCIONES ====================
    elif intencion == 'temporada':
        return """🏖️ <strong>Temporadas de Vuelos en Perú</strong><br><br>

<div style="background: linear-gradient(135deg, #ff4757 0%, #ff6348 100%); color: white; padding: 15px; border-radius: 10px; margin: 10px 0;">
<strong>🔥 TEMPORADA ALTA</strong> (+30-50% más caro)<br>
• 🎄 Diciembre - Febrero: Verano y vacaciones<br>
• 🇵🇪 Julio - Agosto: Fiestas Patrias<br>
• 🐰 Semana Santa (Marzo/Abril)
</div>

<div style="background: linear-gradient(135deg, #ffa502 0%, #ffc048 100%); color: white; padding: 15px; border-radius: 10px; margin: 10px 0;">
<strong>📊 TEMPORADA MEDIA</strong> (Precios normales)<br>
• Junio, Septiembre, Octubre, Noviembre
</div>

<div style="background: linear-gradient(135deg, #26de81 0%, #4ade80 100%); color: white; padding: 15px; border-radius: 10px; margin: 10px 0;">
<strong>✅ TEMPORADA BAJA</strong> (¡Ahorra hasta 40%!)<br>
• Marzo, Abril, Mayo
</div>

<strong>💡 Estrategia ganadora:</strong><br>
Viaja en temporada baja + compra 30-40 días antes = <strong>Máximo ahorro</strong> 💰<br><br>

¿Te gustaría saber algo más? 😊""", False
    
    elif intencion == 'comparar_aerolineas':
        return """✈️ <strong>Comparación de Aerolíneas en Perú</strong><br><br>

<div style="background: #f8f9fa; padding: 12px; border-radius: 8px; margin: 10px 0; border-left: 4px solid #667eea;">
<strong>🔵 LATAM Airlines</strong><br>
✅ Más frecuencias y destinos<br>
✅ Mejor programa de millas (LATAM Pass)<br>
⚠️ Precios 20-30% más altos<br>
✅ Servicio completo incluido<br>
<em>→ Ideal para: Viajes frecuentes, acumular millas</em>
</div>

<div style="background: #f8f9fa; padding: 12px; border-radius: 8px; margin: 10px 0; border-left: 4px solid #ff6348;">
<strong>🔴 Sky Airline</strong><br>
✅ 15-25% más barato que LATAM<br>
⚠️ Menos frecuencias<br>
✅ Equipaje de mano incluido<br>
⚠️ Servicio básico<br>
<em>→ Ideal para: Presupuesto moderado</em>
</div>

<div style="background: #f8f9fa; padding: 12px; border-radius: 8px; margin: 10px 0; border-left: 4px solid #4ade80;">
<strong>🟢 JetSmart</strong><br>
✅ Low-cost, los mejores precios<br>
⚠️ Todo se cobra extra (equipaje, asiento, etc.)<br>
✅ Perfecto para viajes ligeros<br>
⚠️ Menos flexibilidad<br>
<em>→ Ideal para: Máximo ahorro, solo carry-on</em>
</div>

<div style="background: #fff3cd; padding: 12px; border-radius: 8px; border-left: 4px solid #ffc107;">
<strong>⚠️ Importante:</strong> Siempre compara el precio TOTAL (con equipaje y extras incluidos) antes de decidir. A veces la "más barata" termina costando igual.
</div>

¿Quieres que analice cuál te conviene más? 😊""", False
    
    elif intencion == 'cuando_comprar':
        return """⏰ <strong>Guía Completa: Cuándo Comprar Vuelos</strong><br><br>

<div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 18px; border-radius: 12px; margin: 15px 0;">
<strong style="font-size: 18px;">📅 ANTICIPACIÓN ÓPTIMA</strong><br><br>
<div style="background: rgba(255,255,255,0.1); padding: 10px; border-radius: 8px; margin: 8px 0;">
<strong>✅ 30-45 días antes:</strong> ¡PUNTO DULCE! 🎯<br>
<small>El mejor momento para conseguir buenos precios</small>
</div>
<div style="background: rgba(255,255,255,0.1); padding: 10px; border-radius: 8px; margin: 8px 0;">
<strong>💰 15-29 días:</strong> Precios estables<br>
<small>Todavía aceptable, pero menos ofertas</small>
</div>
<div style="background: rgba(255,255,255,0.1); padding: 10px; border-radius: 8px; margin: 8px 0;">
<strong>⚠️ Menos de 15 días:</strong> Precios suben 20-40%<br>
<small>¡Evita comprar tan tarde!</small>
</div>
<div style="background: rgba(255,255,255,0.1); padding: 10px; border-radius: 8px; margin: 8px 0;">
<strong>📈 Más de 60 días:</strong> Precios pueden fluctuar<br>
<small>Aún no se estabilizan</small>
</div>
</div>

<strong>📆 Mejor DÍA para COMPRAR:</strong><br>
<div style="padding: 10px; background: #f0f7ff; border-radius: 8px; margin: 10px 0;">
✅ <strong>Martes y Miércoles:</strong> Mejores ofertas<br>
💰 <strong>Jueves:</strong> Precios moderados<br>
🚫 <strong>Viernes-Domingo:</strong> Más caro (evitar)
</div>

<strong>📆 Mejor DÍA para VIAJAR:</strong><br>
<div style="padding: 10px; background: #f0f7ff; border-radius: 8px; margin: 10px 0;">
✅ <strong>Martes:</strong> Ahorro de 15% vs promedio<br>
✅ <strong>Miércoles:</strong> Ahorro de 12%<br>
⚠️ <strong>Viernes tarde:</strong> Recargo de 25%<br>
🚫 <strong>Domingo tarde:</strong> Recargo de 30%
</div>

<strong>🕐 Mejor HORA para COMPRAR:</strong><br>
<div style="padding: 10px; background: #f0f7ff; border-radius: 8px; margin: 10px 0;">
✅ <strong>10 PM - 2 AM:</strong> Actualizaciones de precios<br>
💰 <strong>Madrugada:</strong> Menos competencia<br>
⚠️ <strong>8 AM - 12 PM:</strong> Precios más altos
</div>

<div style="background: linear-gradient(135deg, #26de81 0%, #4ade80 100%); color: white; padding: 18px; border-radius: 12px; margin: 15px 0;">
<strong style="font-size: 16px;">🎯 FÓRMULA GANADORA:</strong><br><br>
Compra un <strong>MARTES a las 11 PM</strong><br>
Con <strong>30-40 DÍAS</strong> de anticipación<br>
Para volar un <strong>MIÉRCOLES</strong><br>
= <strong>¡AHORRO HASTA 35%!</strong> 💰✨
</div>

¿Quieres más consejos para ahorrar? 😊""", False
    
    elif intencion == 'escalas':
        return """🛫 <strong>Guía Completa sobre Escalas</strong><br><br>

<div style="background: linear-gradient(135deg, #26de81 0%, #4ade80 100%); color: white; padding: 18px; border-radius: 12px; margin: 12px 0;">
<strong style="font-size: 17px;">✈️ VUELO DIRECTO</strong><br><br>
⏱️ <strong>Tiempo:</strong> Ahorra 2-4 horas<br>
💰 <strong>Precio:</strong> +15-30% más caro<br>
✅ <strong>Ventajas:</strong><br>
• Menos cansancio<br>
• Menor riesgo de perder equipaje<br>
• Sin preocupaciones por conexiones<br><br>
<em>→ Ideal para: Viajes de negocios, vuelos cortos, poca flexibilidad</em>
</div>

<div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 18px; border-radius: 12px; margin: 12px 0;">
<strong style="font-size: 17px;">🔄 1 ESCALA</strong><br><br>
⏱️ <strong>Tiempo:</strong> +2-3 horas de viaje total<br>
💰 <strong>Precio:</strong> Balance precio-tiempo<br>
✅ <strong>Ventajas:</strong><br>
• Ahorro moderado<br>
• Tiempo razonable<br>
• Opción de estirar piernas<br>
⚠️ <strong>Riesgos:</strong> Moderados<br><br>
<em>→ Ideal para: Presupuesto moderado, flexibilidad media</em>
</div>

<div style="background: linear-gradient(135deg, #ffa502 0%, #ff6348 100%); color: white; padding: 18px; border-radius: 12px; margin: 12px 0;">
<strong style="font-size: 17px;">🔄🔄 2+ ESCALAS</strong><br><br>
⏱️ <strong>Tiempo:</strong> +4-6 horas (o más)<br>
💰 <strong>Precio:</strong> Hasta 40% más barato<br>
✅ <strong>Ventajas:</strong><br>
• Máximo ahorro<br>
• Posibilidad de conocer ciudades intermedias<br>
⚠️ <strong>Riesgos:</strong> ALTOS<br>
• Mayor probabilidad de retrasos<br>
• Riesgo de perder conexiones<br>
• Mayor desgaste físico<br><br>
<em>→ Ideal para: Presupuesto ajustado, mucha flexibilidad, viajes de placer</em>
</div>

<div style="background: #fff3cd; padding: 15px; border-radius: 10px; border-left: 4px solid #ffc107; margin: 15px 0;">
<strong>💡 RECOMENDACIONES:</strong><br><br>
<strong>Para vuelos DOMÉSTICOS en Perú (1-2h):</strong><br>
→ Prioriza SIEMPRE vuelos directos<br>
→ El ahorro no justifica el tiempo extra<br><br>

<strong>Para vuelos INTERNACIONALES:</strong><br>
→ Evalúa cuánto vale tu tiempo<br>
→ Si ahorras $200 pero pierdes 6 horas, ¿vale la pena?<br><br>

<strong>Si eliges escalas:</strong><br>
✅ Deja MÍNIMO 2 horas entre conexiones<br>
✅ Prefiere misma aerolínea (equipaje directo)<br>
✅ Evita escalas en aeropuertos grandes (más demoras)<br>
⚠️ NO reserves escalas cortas (menos de 90 min)
</div>

<div style="background: #ffe5e5; padding: 15px; border-radius: 10px; border-left: 4px solid #ff4757; margin: 15px 0;">
<strong>⚠️ EVITA ESCALAS SI:</strong><br>
• Viajas con niños pequeños<br>
• Llevas equipaje delicado o importante<br>
• Tienes reuniones/eventos inmediatos al llegar<br>
• El clima puede afectar vuelos (invierno, temporada de huracanes)
</div>

¿Te ayudo a decidir entre directo o con escalas para tu viaje? 😊""", False
    
    elif intencion == 'dias_semana':
        return """📅 <strong>Guía: Mejores Días para Viajar y Ahorrar</strong><br><br>

<div style="background: linear-gradient(135deg, #26de81 0%, #4ade80 100%); color: white; padding: 18px; border-radius: 12px; margin: 15px 0;">
<strong style="font-size: 18px;">✅ DÍAS MÁS BARATOS</strong><br><br>
<div style="background: rgba(255,255,255,0.15); padding: 12px; border-radius: 8px; margin: 8px 0;">
<strong>🔵 MARTES</strong><br>
Ahorro: <strong>-15%</strong> vs promedio<br>
<small>El mejor día de toda la semana</small>
</div>
<div style="background: rgba(255,255,255,0.15); padding: 12px; border-radius: 8px; margin: 8px 0;">
<strong>🔵 MIÉRCOLES</strong><br>
Ahorro: <strong>-12%</strong> vs promedio<br>
<small>Segundo mejor día</small>
</div>
<div style="background: rgba(255,255,255,0.15); padding: 12px; border-radius: 8px; margin: 8px 0;">
<strong>🔵 SÁBADO (mediodía)</strong><br>
Ahorro: <strong>-8%</strong> vs promedio<br>
<small>Buena opción de fin de semana</small>
</div>
</div>

<div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 18px; border-radius: 12px; margin: 15px 0;">
<strong style="font-size: 18px;">💰 PRECIO NORMAL</strong><br><br>
<strong>⚪ LUNES:</strong> Precio estándar<br>
<strong>⚪ JUEVES:</strong> Precio estándar<br>
<strong>⚪ SÁBADO (mañana):</strong> Precio estándar
</div>

<div style="background: linear-gradient(135deg, #ff4757 0%, #ff6348 100%); color: white; padding: 18px; border-radius: 12px; margin: 15px 0;">
<strong style="font-size: 18px;">🚫 DÍAS MÁS CAROS</strong><br><br>
<div style="background: rgba(255,255,255,0.15); padding: 12px; border-radius: 8px; margin: 8px 0;">
<strong>🔴 VIERNES (tarde/noche)</strong><br>
Recargo: <strong>+20-30%</strong><br>
<small>Inicio de fin de semana laboral</small>
</div>
<div style="background: rgba(255,255,255,0.15); padding: 12px; border-radius: 8px; margin: 8px 0;">
<strong>🔴 DOMINGO (tarde/noche)</strong><br>
Recargo: <strong>+25-35%</strong><br>
<small>Retorno de fin de semana - EL MÁS CARO</small>
</div>
<div style="background: rgba(255,255,255,0.15); padding: 12px; border-radius: 8px; margin: 8px 0;">
<strong>🔴 LUNES (mañana temprano)</strong><br>
Recargo: <strong>+15-20%</strong><br>
<small>Viajes de negocios</small>
</div>
</div>

<div style="background: #f0f7ff; padding: 15px; border-radius: 10px; border-left: 4px solid #667eea; margin: 15px 0;">
<strong>📊 ¿POR QUÉ SUCEDE ESTO?</strong><br><br>
<strong>Viernes PM:</strong> Todos salen de viaje de fin de semana<br>
<strong>Domingo PM:</strong> Todos regresan a casa/trabajo<br>
<strong>Lunes AM:</strong> Viajes de negocios concentrados<br>
<strong>Martes-Miércoles:</strong> Baja demanda = mejores precios
</div>

<div style="background: linear-gradient(135deg, #26de81 0%, #4ade80 100%); color: white; padding: 18px; border-radius: 12px; margin: 15px 0;">
<strong style="font-size: 16px;">🎯 ESTRATEGIA MAESTRA:</strong><br><br>
1️⃣ Viaja <strong>MARTES</strong> o <strong>MIÉRCOLES</strong><br>
2️⃣ Compra el boleto un <strong>MARTES</strong> por la noche<br>
3️⃣ Con <strong>30-40 DÍAS</strong> de anticipación<br>
4️⃣ En <strong>TEMPORADA BAJA</strong> (Mar-May)<br><br>
= <strong>¡AHORRO TOTAL: HASTA 50%!</strong> 💰✨
</div>

<div style="background: #fff3cd; padding: 15px; border-radius: 10px; border-left: 4px solid #ffc107; margin: 15px 0;">
<strong>💡 TIP EXTRA:</strong><br>
Si tu trabajo lo permite, toma vacaciones martes a jueves en lugar de viernes a domingo. Podrías ahorrar cientos de soles solo cambiando días.
</div>

¿Quieres saber algo más sobre cómo ahorrar? 😊""", False
    
    elif intencion == 'consejos':
        return """💡 <strong>Guía Definitiva: Tips PRO para Ahorrar en Vuelos</strong><br><br>

<div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 18px; border-radius: 12px; margin: 15px 0;">
<strong style="font-size: 18px;">🎯 LOS 10 SECRETOS MÁS IMPORTANTES</strong>
</div>

<div style="background: #f8f9fa; padding: 15px; border-radius: 10px; margin: 10px 0; border-left: 4px solid #667eea;">
<strong>1️⃣ Usa SIEMPRE Modo Incógnito</strong><br>
Las aerolíneas rastrean tus búsquedas con cookies y suben los precios cada vez que vuelves a buscar el mismo vuelo. 🕵️<br>
<em>→ Ahorro potencial: 10-15%</em>
</div>

<div style="background: #f8f9fa; padding: 15px; border-radius: 10px; margin: 10px 0; border-left: 4px solid #4ade80;">
<strong>2️⃣ Configura Alertas de Precio</strong><br>
Google Flights, Skyscanner o Kayak te avisan cuando bajan los precios de tu ruta. 📧<br>
<em>→ No pierdas ofertas flash</em>
</div>

<div style="background: #f8f9fa; padding: 15px; border-radius: 10px; margin: 10px 0; border-left: 4px solid #ff6348;">
<strong>3️⃣ Flexibilidad de ±3 Días</strong><br>
Si puedes mover tu viaje 3 días antes o después, ahorras hasta 30%. Usa calendarios de precios. 📅<br>
<em>→ Ahorro potencial: 20-30%</em>
</div>

<div style="background: #f8f9fa; padding: 15px; border-radius: 10px; margin: 10px 0; border-left: 4px solid #ffa502;">
<strong>4️⃣ Aeropuertos Alternativos</strong><br>
A veces volar desde/hacia ciudades cercanas es más barato. Ejemplo: Callao vs Centro Lima. 🛫<br>
<em>→ Ahorro potencial: 15-25%</em>
</div>

<div style="background: #f8f9fa; padding: 15px; border-radius: 10px; margin: 10px 0; border-left: 4px solid #26de81;">
<strong>5️⃣ Suscríbete a Newsletters</strong><br>
LATAM, Sky y JetSmart envían ofertas flash EXCLUSIVAS a suscriptores antes que al público. 📬<br>
<em>→ Acceso a ofertas limitadas</em>
</div>

<div style="background: #f8f9fa; padding: 15px; border-radius: 10px; margin: 10px 0; border-left: 4px solid #764ba2;">
<strong>6️⃣ Acumula Millas</strong><br>
Incluso en vuelos económicos, acumula puntos. LATAM Pass es el más útil en Perú. ✈️<br>
<em>→ Vuelos gratis a largo plazo</em>
</div>

<div style="background: #f8f9fa; padding: 15px; border-radius: 10px; margin: 10px 0; border-left: 4px solid #667eea;">
<strong>7️⃣ Viaja Solo con Carry-on</strong><br>
Evita costos de equipaje documentado. Ahorras dinero y tiempo en el aeropuerto. 🎒<br>
<em>→ Ahorro: S/80-150 por vuelo</em>
</div>

<div style="background: #f8f9fa; padding: 15px; border-radius: 10px; margin: 10px 0; border-left: 4px solid #ffc107;">
<strong>8️⃣ Compara Monedas</strong><br>
A veces pagar en soles vs dólares hace diferencia. Prueba ambas opciones. 💱<br>
<em>→ Ahorro potencial: 5-10%</em>
</div>

<div style="background: #f8f9fa; padding: 15px; border-radius: 10px; margin: 10px 0; border-left: 4px solid #ff4757;">
<strong>9️⃣ Evita Fines de Semana</strong><br>
Comprar y viajar entre semana es SIEMPRE más barato. 📆<br>
<em>→ Ahorro potencial: 25-35%</em>
</div>

<div style="background: #f8f9fa; padding: 15px; border-radius: 10px; margin: 10px 0; border-left: 4px solid #4ade80;">
<strong>🔟 Usa Tarjetas con Beneficios</strong><br>
Algunas tarjetas de crédito ofrecen millas, seguros de viaje o descuentos. 💳<br>
<em>→ Beneficios adicionales gratis</em>
</div>

<div style="background: linear-gradient(135deg, #26de81 0%, #4ade80 100%); color: white; padding: 20px; border-radius: 12px; margin: 20px 0;">
<strong style="font-size: 18px;">🏆 FÓRMULA MAESTRA DEFINITIVA</strong><br><br>
<div style="background: rgba(255,255,255,0.15); padding: 15px; border-radius: 8px;">
✅ Modo incógnito<br>
✅ Compra un MARTES a las 11 PM<br>
✅ Con 30-40 días de anticipación<br>
✅ Para volar un MIÉRCOLES<br>
✅ En temporada BAJA (Mar-May)<br>
✅ Solo con carry-on<br>
✅ Compara 3 aerolíneas<br><br>
= <strong>¡AHORRO MÁXIMO POSIBLE: 40-50%!</strong> 💰🎉
</div>
</div>

<div style="background: #fff3cd; padding: 15px; border-radius: 10px; border-left: 4px solid #ffc107; margin: 15px 0;">
<strong>⚡ BONUS TIP:</strong><br>
Si ves un buen precio, NO lo pienses mucho. Los algoritmos de aerolíneas detectan cuando muchas personas buscan la misma ruta y suben los precios en minutos. ¡Actúa rápido!
</div>

¿Quieres que analice tu predicción con estos consejos en mente? 😊""", False
    
    # ==================== RESPUESTA GENÉRICA ====================
    else:
        return """👋 <strong>Hola, estoy aquí para ayudarte</strong><br><br>

Puedo responder preguntas sobre:<br><br>

<div style="display: grid; gap: 10px;">
<div style="background: #f0f7ff; padding: 12px; border-radius: 8px; border-left: 4px solid #667eea;">
📊 <strong>"Analiza mi predicción"</strong><br>
<small>Te doy un análisis completo y personalizado</small>
</div>

<div style="background: #f0f7ff; padding: 12px; border-radius: 8px; border-left: 4px solid #ff6348;">
🏖️ <strong>"¿Cuándo es temporada alta?"</strong><br>
<small>Mejores fechas para viajar y ahorrar</small>
</div>

<div style="background: #f0f7ff; padding: 12px; border-radius: 8px; border-left: 4px solid #26de81;">
⏰ <strong>"¿Cuál es el mejor momento?"</strong><br>
<small>Estrategias para conseguir mejor precio</small>
</div>

<div style="background: #f0f7ff; padding: 12px; border-radius: 8px; border-left: 4px solid #ffa502;">
✈️ <strong>"Compara aerolíneas"</strong><br>
<small>LATAM vs Sky vs JetSmart</small>
</div>

<div style="background: #f0f7ff; padding: 12px; border-radius: 8px; border-left: 4px solid #764ba2;">
💡 <strong>"Dame consejos"</strong><br>
<small>Tips profesionales para ahorrar</small>
</div>
</div>

<br>También puedo ayudarte con:<br>
• Información sobre escalas<br>
• Mejores días para viajar<br>
• Preguntas específicas sobre tu viaje<br><br>

¿En qué te puedo ayudar hoy? 😊""", False    

# ========== MANEJO DE ERRORES ==========
@app.errorhandler(404)
def no_encontrado(error):
    try:
        return render_template('error.html',
                             mensaje='Página no encontrada',
                             detalle='La página que buscas no existe'), 404
    except:
        return '''
        <html>
        <body style="font-family: Arial; text-align: center; padding: 50px;">
            <h1>404 - Página no encontrada</h1>
            <p>La página que buscas no existe</p>
            <a href="/">Volver al inicio</a>
        </body>
        </html>
        ''', 404

@app.errorhandler(500)
def error_interno(error):
    try:
        return render_template('error.html',
                             mensaje='Error interno',
                             detalle='Ocurrió un error en el servidor'), 500
    except:
        return '''
        <html>
        <body style="font-family: Arial; text-align: center; padding: 50px;">
            <h1>500 - Error interno</h1>
            <p>Ocurrió un error en el servidor</p>
            <a href="/">Volver al inicio</a>
        </body>
        </html>
        ''', 500
# ========== MANEJO DE ERRORES ==========
#@app.errorhandler(404)
#def no_encontrado(error):
#    return render_template('error.html',
#                         mensaje='Página no encontrada',
#                         detalle='La página que buscas no existe'), 404
#
#@app.errorhandler(500)
#def error_interno(error):
#    return render_template('error.html',
#                         mensaje='Error interno',
#                         detalle='Ocurrió un error en el servidor'), 500

# ========== INICIALIZACIÓN1 ==========
#if __name__ == '__main__':
#    with app.app_context():
#        db.create_all()  # Crear tablas si no existen
    
#   print("🚀 Iniciando aplicación Flask...")
    
#    if cargar_modelo():
#        print("✓ Modelo cargado exitosamente")
#    else:
#        print("⚠️  Modelo no encontrado")
#    
#    if cargar_datos_cache():
#        print("✓ Datos cargados en caché")
#    else:
#        print("⚠️  Datos no encontrados")
    
    #app.run(debug=True, host='0.0.0.0', port=5000)
    # CAMBIO AQUÍ ⬇️
#    port = int(os.environ.get('PORT', 5000))
#    debug = os.environ.get('FLASK_ENV') != 'production'
#    app.run(debug=debug, host='0.0.0.0', port=port)
# ========== INICIALIZACIÓN ==========
#if __name__ == '__main__':
#    with app.app_context():
#        try:
#            db.create_all()  # Crear tablas si no existen
#            print("✓ Tablas de base de datos creadas/verificadas")
#        except Exception as e:
#            print(f"⚠️ Error creando tablas: {e}")
#    
#    print("🚀 Iniciando aplicación Flask...")
#    
#    if cargar_modelo():
#        print("✓ Modelo cargado exitosamente")
#    else:
#        print("⚠️ Modelo no encontrado - Ejecuta training.py primero")
#    
#    if cargar_datos_cache():
#        print("✓ Datos cargados en caché")
#    else:
#        print("⚠️ Datos no encontrados - Ejecuta generar_datos.py primero")
#    
#    # Configuración para producción
#    port = int(os.environ.get('PORT', 5000))
#    debug = os.environ.get('FLASK_ENV') != 'production'
#    app.run(debug=debug, host='0.0.0.0', port=port) 

# ========== INICIALIZACIÓN ==========
if __name__ == '__main__':
    with app.app_context():
        try:
            db.create_all()
            print("✓ Tablas de base de datos creadas/verificadas")
        except Exception as e:
            print(f"⚠️ Error creando tablas: {e}")
    
    print("🚀 Iniciando aplicación Flask...")
    
    # Cargar datos primero (necesarios para el modelo)
    if cargar_datos_cache():
        print("✓ Datos disponibles")
    else:
        print("⚠️ No se pudieron cargar datos")
    
    # Luego cargar modelo
    if cargar_modelo():
        print("✓ Modelo disponible")
    else:
        print("⚠️ Modelo no disponible")
    
    # Configuración para producción
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') != 'production'
    app.run(debug=debug, host='0.0.0.0', port=port)

# ========== INICIALIZACIÓN PARA GUNICORN ==========
# Esto se ejecuta cuando gunicorn carga la app
print("🔧 Inicializando app para Gunicorn...")

# Crear tablas
with app.app_context():
    try:
        db.create_all()
        print("✓ Base de datos inicializada")
    except Exception as e:
        print(f"⚠️ Error en base de datos: {e}")

# Cargar datos y modelo al inicio
print("📊 Cargando recursos...")
cargar_datos_cache()
cargar_modelo()
print("✅ App lista para recibir peticiones")